from selenium import webdriver
import os
import sys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
caminho_absoluto = os.path.abspath(os.curdir)
sys.path.insert(0, caminho_absoluto)
from Senhas.login_hapvida import realizarLogin
import time
import json
import shutil
import openpyxl
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
#Caso Edge:
from selenium.webdriver.edge.options import Options



# MENU PRINCIPAL 
print('===== BEM VINDO AO GSYSTEM =====:\n')
print('===== BAIXA REFERENTE AOS BOLETOS HAPVIDA =====:\n')
print('Responsável: Giovanni.Souza')
input('CONFIRMAÇAO DE SENHA\n\n')

opc_opr = input("INFORME A OPERADORA:\n[1]Affix\n[2]Alter\n")
if opc_opr == "1":
    operadora = "AFFIX"
    aba_excel = "Hapvida Affix"

elif opc_opr == "2":
    operadora = "ALTER"
    aba_excel = "Hapvida Alter"


opc_tipo_contrato = input("INFORME O TIPO DE CONTRATO:\n[1]MENSALIDADE\n[2]COPART\n")
if opc_tipo_contrato == "1":
    #O que vai na descrição
    tipo_contrato = ""
    #O que aparece no Xpath
    tipo_contrato_site = "Coparticipação"

elif opc_tipo_contrato == "2":
    tipo_contrato = "MENSALIDADE"
    tipo_contrato_site = "Mensalidade" 


print("INFORME A DATA DE VENCIMENTO DESEJADA SEGUIDA DO DIA, MES E ANO:\n")
dia_venc = input('Dia(DD): ')
mes_venc = input('Mes(MM): ')
ano_venc = input('Ano(AAAA): ')

#formato data venc
data_completa_venc = f"{dia_venc}/{mes_venc}/{ano_venc}"


#lista de contratos
"""lista_contratos = [
    "BNIZZ",
    "FYR8W",
    "RT5QU"
]"""


# Carregr arquivo
caminho_arquivo = 'Senhas/senhas_operadoras.xlsx'
wb = openpyxl.load_workbook(caminho_arquivo)
aba_hapvida = wb["teste"]

for row in range(1, aba_hapvida.max_row + 1):
    contrato = str(aba_hapvida.cell(row=row, column=1).value)
    if len(str(aba_hapvida[f'G{row}'].value)) == 1:
        plan_venc = str(f"0{aba_hapvida[f'G{row}'].value}")
        plan_admn = str(aba_hapvida[f'C{row}'].value)
        plan_admn_edit = plan_admn.replace(" ", "")
    else:
        #plan = planilha
        plan_venc = str(aba_hapvida[f'G{row}'].value)
        plan_admn = str(aba_hapvida[f'C{row}'].value)
        plan_admn_edit = plan_admn.replace(" ", "")

    if dia_venc == plan_venc and plan_admn_edit == operadora:



        # CONFIGURACOES PADROES:
        # Ajuste de caminho para importar funções personalizadas
        caminho_absoluto = os.path.abspath(os.curdir)

        # Identificacao de pastas na arquitetura do projeto (includes)
        sys.path.insert(0, caminho_absoluto)

        
        # CONFIGURAÇÃO DE IMPRESSÃO (Obrigatório vir antes de iniciar o Chrome)
        #PREFERENCIA NAVEGADOR
        #chrome_options = Options()
        #chrome_options.add_argument('--kiosk-printing') # Ativa a impressão sem perguntas
        edge_options = Options()
        edge_options.add_argument("--kiosk-printing")
        
        

        # Configura o destino como "Salvar como PDF" e cria pasta download na pasta projeto
        pasta_download = os.path.join(os.getcwd(), "Download")
        os.makedirs(pasta_download, exist_ok=True)

        prefs = {
            "printing.print_preview_sticky_settings.appState": json.dumps({
                "recentDestinations": [{"id": "Save as PDF", "origin": "local", "account": ""}],
                "selectedDestinationId": "Save as PDF",
                "version": 2
            }),
            "savefile.default_directory": pasta_download
        }
        
        
        #PREFERENCIA NAVEGADOR
        #chrome_options.add_experimental_option('prefs', prefs)
        edge_options.add_experimental_option('prefs', prefs)
        

        # INSTANCIANDO O NAVEGADOR COM AS OPÇÕES
        #PREFERENCIA NAVEGADOR
        #navegador = webdriver.Chrome(options=chrome_options)
        navegador = webdriver.Edge(options=edge_options)



        # DEFININDO VARIAVEIS
        tipo_arquivo = "hapvida_boleto"



        # FLUXO DE NAVEGACAO

        # Acessar link boleto hapvida
        navegador.get("https://webhap.hapvida.com.br/pls/webhap/webNewBoletoEmpresa.Login")

        # Tempo para carregar a página
        time.sleep(5)

        # Fechar pop-up tela de login
        botao = navegador.find_element(By.CLASS_NAME, "ui-button-text")
        botao.click()

        time.sleep(1)
        navegador.maximize_window()

        # REALIZA LOGIN USANDO FUNCAO DE FORA (realizarLogin())

        #lista de contratos

        realizarLogin(navegador, contrato, tipo_arquivo, "entidade", "cnpj")

        try:
            #Aguarda a tela do iframe entrar
            iframe_faturas = WebDriverWait(navegador, 20).until(
                EC.presence_of_element_located((By.TAG_NAME, "iframe"))
            )
            
            # 2. Encontrou iframe, usar o switch_to.frame para acessá-lo
            # para caso precisar sair do iframe: navegador.switch_to.default_content()
            navegador.switch_to.frame(iframe_faturas)
  
        
            time.sleep(2)
            
            # 3. Localizar o primeiro checkbox
            xpath_checkbox_mestre = "//div[contains(@class, 'tableHeaderCell')]//button[@role='checkbox']"
            
            checkbox_pai = WebDriverWait(navegador, 15).until(
                EC.presence_of_element_located((By.XPATH, xpath_checkbox_mestre))
            )
            
            # 4. Forçar com javascript o clique
            navegador.execute_script("arguments[0].click();", checkbox_pai)


        except Exception as e:
            print(f"Erro durante a execução: {e}")
            
        wait = WebDriverWait(navegador, 10)
        linhas = wait.until(
            EC.presence_of_all_elements_located((By.XPATH, "//div[contains(@class, '__tableRow')]"))
        )
        for linha in linhas:
            
            try: 
                valida_tipo_contrato = linha.find_element(
                By.XPATH, "./div[2][contains(string(), 'Coparticipação')]"
                )

                valida_venc_contrato = linha.find_element(
                By.XPATH, f"./div[3][contains(string(), '{data_completa_venc}')]"
                )

                print(valida_tipo_contrato.text)
                print(valida_venc_contrato.text)

                """linha.find_element(By.XPATH, ".//button")
                # Seleciona o boleto de acordo com vencimento
                time.sleep(4) # Espera carregar após login


                # Muda para a aba onde o PDF é exibido (Aba 1)
                time.sleep(3)
                aba_final = navegador.window_handles
                navegador.switch_to.window(aba_final[1])

                # --- COMANDO FINAL PARA SALVAR ---
                # Como foi ativado o Modo Kiosk, ele salva o PDF na pasta sem abrir a janela de impressão
                time.sleep(2)
                navegador.execute_script("window.print();")


                # --- RENOMEANDO ARQUIVO DA PASTA DOWNLOAD ENCAMINHANDO PARA PASTA DA OPERADORA ---

                # Lista todos os arquivos PDF na pasta

                pdfs = []

                for f in os.listdir(pasta_download):
                    if f.endswith(".pdf"):
                        pdfs.append(f)

                # Se houver pelo menos um PDF, renomeia o primeiro
                if pdfs:
                    arquivo_antigo = os.path.join(pasta_download, pdfs[0])
                    arquivo_novo = os.path.join(pasta_download, f"{dia_venc}.{mes_venc}.{ano_venc}-{operadora} HAPVIDA {contrato} Boleto {tipo_contrato}.pdf")
                    os.rename(arquivo_antigo, arquivo_novo)
                    #print("Arquivo renomeado com sucesso!")
                else:
                    print("Nenhum arquivo PDF encontrado.")
                    exit()


                # Mover arquivo renomeado para a pasta de vencimento da operadora (Hapvida Arquivo)
                # Caso nao existir, cria
                pasta_destino = os.path.join("Affix", "Operadoras", "Hapvida", "Mensalidade_Copart", "Hapvida Arquivo",f"{dia_venc}_{mes_venc}_{ano_venc}") 
                os.makedirs(pasta_destino, exist_ok=True)

                shutil.move(arquivo_novo, pasta_destino)
                print(f"BAIXADO: {arquivo_novo}")
                # Fim da operacao
                navegador.quit()
                time.sleep(2)"""



            except NoSuchElementException:
                continue
            
            






                