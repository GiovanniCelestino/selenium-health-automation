from selenium import webdriver
import os
import sys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
caminho_absoluto = os.path.abspath(os.curdir)
sys.path.insert(0, caminho_absoluto)
from senhas.login_hapvida import realizarLogin
import time
import json
import shutil
#Caso Edge:
from selenium.webdriver.edge.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
import openpyxl
#Excessoes Try/Except
from selenium.common.exceptions import NoSuchElementException
from scripts.imprime_arquivos.config_impressao import *




# Função para baixa de arquivo .CSV ou .PDF
def baixa_arquivo_extensao(navegador, linha_atual, texto_da_linha, contrato):
    if "CSV" in texto_da_linha:
  
    
        arquivo_csv = linha_atual.find_element(By.XPATH, "./td[1]/small/a")
        arquivo_csv.click()
        time.sleep(2)
        baixar_csv = navegador.find_element(By.XPATH, ".//a[text()='clique aqui']")
        baixar_csv.click()
        navegador.find_element(By.XPATH, ".//a[text()='Voltar']").click()
        campo_pesquisa = navegador.find_element(By.XPATH, "//input[@type='search']")
        campo_pesquisa.clear()
        campo_pesquisa.send_keys(contrato)
        extencao = "CSV"
        return extencao
        

    if "PDF" in texto_da_linha:
     
        
        arquivo_pdf = linha_atual.find_element(By.XPATH, "./td[1]/small/a")
        arquivo_pdf.click()
        time.sleep(2)
        baixar_pdf = navegador.find_element(By.XPATH, ".//a[text()='clique aqui']")
        baixar_pdf.click()
        navegador.find_element(By.XPATH, ".//a[text()='Voltar']").click()
        campo_pesquisa = navegador.find_element(By.XPATH, "//input[@type='search']")
        campo_pesquisa.clear()
        time.sleep(2)
        campo_pesquisa.send_keys(contrato)
        extencao = "PDF"
        return extencao






# MENU PRINCIPAL 
print('===== BEM VINDO AO GSYSTEM =====:\n')
print('===== BAIXA REFERENTE AOS BOLETOS HAPVIDA =====:\n')
print('Responsável: Giovanni.Souza')

opc_opr = input("INFORME A OPERADORA:\n[1]Affix\n[2]Alter\n")
if opc_opr == "1":
    operadora = "AFFIX"
    aba_excel = "Hapvida Affix"

elif opc_opr == "2":
    operadora = "ALTER"
    aba_excel = "Hapvida Alter"


opc_tipo_contrato = input("INFORME O TIPO DE CONTRATO:\n[1]MENSALIDADE\n[2]COPART\n")
if opc_tipo_contrato == "2":
    #O que vai na descrição
    tipo_contrato = "COPART"
    tipo_contrato_cond = "S"

elif opc_tipo_contrato == "1":
    tipo_contrato = "MENSALIDADE"
    tipo_contrato_cond = "N"


print("INFORME A DATA DE VENCIMENTO DESEJADA SEGUIDA DO DIA, MES E ANO:\n")
dia_venc = input('Dia(DD): ')
mes_venc = input('Mes(MM): ')
ano_venc = input('Ano(AAAA): ')


#formato data venc
data_completa_venc = f"{dia_venc}/{mes_venc}/{ano_venc}"
#converter em data
data_completa_convert = datetime.strptime(data_completa_venc, "%d/%m/%Y").date()


#data atual
data_hoje = datetime.now().date()

print(data_completa_convert)
print(data_hoje)




# CONFIGURACOES PADROES:
# Ajuste de caminho para importar funções personalizadas
caminho_absoluto = os.path.abspath(os.curdir)

# Identificacao de pastas na arquitetura do projeto (includes)
sys.path.insert(0, caminho_absoluto)

config_edge, pasta_download = ativa_impressao()
# INSTANCIANDO O NAVEGADOR COM AS OPÇÕES
#PREFERENCIA NAVEGADOR
#navegador = webdriver.Chrome(options=chrome_options)
navegador = webdriver.Edge(options=config_edge)


# DEFININDO VARIAVEIS

tipo_arquivo = "hapvida_detalhe_affix"
cnpj = "Indefinido"




# REALIZA LOGIN USANDO FUNCAO DE FORA (realizarLogin())
#lista de contratos

# Carregr arquivo
caminho_arquivo = 'senhas/senhas_operadoras.xlsx'
wb = openpyxl.load_workbook(caminho_arquivo)
aba_hapvida = wb["teste"]

        # MANIPULAÇÃO PLANILHA

for row in range(1, aba_hapvida.max_row + 1):
    contrato = str(aba_hapvida.cell(row=row, column=1).value)
    #input(f"Valor Celula:{valor_celula}\nContrato Procurado:{contrato_procurado}")
    
    #Checa se vencimento tem 1 caractere. Caso sim, acrescenta 0 na frente
    if len(str(aba_hapvida[f'G{row}'].value)) == 1:
        plan_venc = str(f"0{aba_hapvida[f'G{row}'].value}")
        plan_admn = str(aba_hapvida[f'C{row}'].value)
        plan_admn_edit = plan_admn.replace(" ", "")
        plan_tipo_cont = str(f"{aba_hapvida[f'F{row}'].value}")

    else:
        plan_venc = str(aba_hapvida[f'G{row}'].value)
        plan_admn = str(aba_hapvida[f'C{row}'].value)
        plan_admn_edit = plan_admn.replace(" ", "")
        plan_tipo_cont = str(aba_hapvida[f'F{row}'].value)

    
    if dia_venc == plan_venc and plan_admn_edit == "AFFIX" and tipo_contrato_cond == plan_tipo_cont:
        cnpj = aba_hapvida[f'D{row}'].value
        
        
        # FLUXO DE NAVEGACAO
        # Acessar link nota fiscal hapvida affix
        navegador.get("https://www.hapvida.com.br/pls/webhap/webNewTrocaArquivo.login")

        # Tempo para carregar a página
        time.sleep(3)

        # Maximixar página
        navegador.maximize_window()
        
        realizarLogin(navegador, contrato, tipo_arquivo, "entidade", cnpj)

        if data_completa_convert <= data_hoje:
            arquivo_anterior = navegador.find_element(By.XPATH, "//a[text()='BAIXAR ARQUIVOS - DOWNLOAD (Anterior)']")
            arquivo_anterior.click()



            campo_pesquisa = navegador.find_element(By.XPATH, "//input[@type='search']")
                
            #limpa registros do campo 
            campo_pesquisa.clear()

            campo_pesquisa.send_keys(contrato)
            #procurar pelo contrato específico:
           
            
            return_arquivo_csv = False
            return_arquivo_pdf = False
            #id_proximo = navegador.find_element(By.ID, "table_id_next")
            #classe_proximo = id_proximo.get_attribute("class")
            continuar_paginacao = True
            while continuar_paginacao:
                total_linhas = len(navegador.find_elements(By.XPATH, "//table[@id='table_id']/tbody/tr"))

                for indice in range(1, total_linhas + 1):
                    
                    
                    try:
                        linha_atual = navegador.find_element(By.XPATH, f"//table[@id='table_id']/tbody/tr[{indice}]")
                    except NoSuchElementException:
                        break

                    # Checa se ambos os arquivos já foram baixados
                    if return_arquivo_csv and return_arquivo_pdf:
                        continuar_paginacao = False
                        break
                    
                    
                    texto_da_linha = linha_atual.text

                    try:
                        id_proximo = navegador.find_element(By.ID, "table_id_next")
                        classe_proximo = id_proximo.get_attribute("class")
                        
                        if indice == total_linhas:
                            if classe_proximo == "paginate_button next":
                                navegador.find_element(By.ID, "table_id_next").click()
                                time.sleep(3) 
                                break
                            elif classe_proximo == "paginate_button next disabled":
                                continuar_paginacao = False
                                break

                        # Tenta pegar o período da linha atual
                        observacao = linha_atual.find_element(By.XPATH, "./td[4]").text  
                        data_periodo_contrato = linha_atual.find_element(By.XPATH, "./td[2]").text   
                        data_inicio = data_periodo_contrato[0:10]
                        data_fim = data_periodo_contrato[15:25]

                        if "/" in data_fim and data_inicio: 
                            data_inicio_convert = datetime.strptime(data_inicio, "%d/%m/%Y").date()
                            data_fim_convert = datetime.strptime(data_fim, "%d/%m/%Y").date()

                        if tipo_contrato_cond == "S":
                            if (data_completa_convert >= data_inicio_convert and data_completa_convert <= data_fim_convert) and "Remessa" in observacao:
                                return_extensao = baixa_arquivo_extensao(navegador, linha_atual, texto_da_linha, contrato)
                                if return_extensao == "PDF":
                                    return_arquivo_pdf = True
                                elif return_extensao == "CSV":
                                    return_arquivo_csv = True

                        elif tipo_contrato_cond == "N":
                            if (data_completa_convert >= data_inicio_convert and data_completa_convert <= data_fim_convert) and "Remessa" not in observacao:
                                return_extensao = baixa_arquivo_extensao(navegador, linha_atual, texto_da_linha, contrato)
                                if return_extensao == "PDF":
                                    return_arquivo_pdf = True
                                elif return_extensao == "CSV":
                                    return_arquivo_csv = True

                    except NoSuchElementException:
                        print("Linha ignorada (não possui a coluna de período ou mudou de estrutura).")
                        continuar_paginacao = False
                        break

