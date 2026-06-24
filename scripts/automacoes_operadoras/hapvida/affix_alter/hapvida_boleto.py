from selenium import webdriver
import os
import sys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
caminho_absoluto = os.path.abspath(os.curdir)
sys.path.insert(0, caminho_absoluto)
from senhas.login_hapvida import realizarLogin
import time
import json
import shutil
import openpyxl
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
#Caso Edge:
from selenium.webdriver.edge.options import Options
from pathlib import Path
from selenium.common.exceptions import TimeoutException
from scripts.imprime_arquivos.config_impressao import *



# MENU PRINCIPAL 
print('===== BEM VINDO AO GSYSTEM =====:\n')
print('===== BAIXA REFERENTE AOS BOLETOS HAPVIDA =====:\n')
print('Responsável: Giovanni.Souza')

opc_opr = input("INFORME A OPERADORA:\n[1]Affix\n[2]Alter\n")
if opc_opr == "1":
    operadora = "AFFIX"
    aba_excel = "Hapvida Affix"

elif opc_opr == "2":
    operadora = "ALTER"
    aba_excel = "Hapvida Alter"


opc_tipo_contrato = input("INFORME O TIPO DE CONTRATO:\n[1]MENSALIDADE\n[2]COPART\n")
if opc_tipo_contrato == "2":
    #O que vai na descrição
    tipo_contrato = "COPART"
    #O que aparece no Xpath
    tipo_contrato_site = "Coparticipação"
    tipo_contrato_cond = "S"

elif opc_tipo_contrato == "1":
    tipo_contrato = "MENSALIDADE"
    tipo_contrato_site = "Mensalidade" 
    tipo_contrato_cond = "N"


print("INFORME A DATA DE VENCIMENTO DESEJADA SEGUIDA DO DIA, MES E ANO:\n")
dia_venc = input('Dia(DD): ')
mes_venc = input('Mes(MM): ')
ano_venc = input('Ano(AAAA): ')

#formato data venc
data_completa_venc = f"{dia_venc}/{mes_venc}/{ano_venc}"


# Carregar arquivo
caminho_arquivo = 'senhas/senhas_operadoras.xlsx'
wb = openpyxl.load_workbook(caminho_arquivo)
aba_hapvida = wb[aba_excel]

for row in range(1, aba_hapvida.max_row + 1):
    
    contrato = str(aba_hapvida.cell(row=row, column=1).value)
    
    if len(str(aba_hapvida[f'G{row}'].value)) == 1:
        plan_venc = str(f"0{aba_hapvida[f'G{row}'].value}")
        plan_admn = str(aba_hapvida[f'C{row}'].value)
        plan_admn_edit = plan_admn.replace(" ", "")
        plan_tipo_cont = str(aba_hapvida[f'F{row}'].value)
    else:
        #plan = planilha
        plan_venc = str(aba_hapvida[f'G{row}'].value)
        plan_admn = str(aba_hapvida[f'C{row}'].value)
        plan_admn_edit = plan_admn.replace(" ", "")
        plan_tipo_cont = str(aba_hapvida[f'F{row}'].value)

    if dia_venc == plan_venc and plan_admn_edit == operadora and tipo_contrato_cond == plan_tipo_cont:
        data_download = f"{dia_venc}.{mes_venc}.{ano_venc}"
        #Verifica se existe o contrato na pasta
        pasta_arquivos_baixados = os.path.join(os.getcwd(), f"dados\\arquivos_direcionados\\operadoras\\hapvida_arquivos\\affix_alter_arquivos\\{dia_venc}_{mes_venc}_{ano_venc}")
        
        #Verifica se existe a pasta de destino, caso contrário cria.
        pasta_destino = os.path.join("dados", "arquivos_direcionados", "operadoras", "hapvida_arquivos", "affix_alter_arquivos",f"{dia_venc}_{mes_venc}_{ano_venc}") 
        os.makedirs(pasta_destino, exist_ok=True)
        
        #Checa se o contrato já foi baixado
        for f in os.listdir(pasta_arquivos_baixados):
            if data_download in f and contrato in f and tipo_contrato in f:
                print(f'Contrato {contrato} já foi baixado.')
                break
        else: 

            # CONFIGURACOES PADROES:
            # Ajuste de caminho para importar funções personalizadas
            caminho_absoluto = os.path.abspath(os.curdir)

            # Identificacao de pastas na arquitetura do projeto (includes)
            sys.path.insert(0, caminho_absoluto)

            # Chama função para configuração de impressão
            config_edge, pasta_download = ativa_impressao()

            # INSTANCIANDO O NAVEGADOR COM AS OPÇÕES
            #PREFERENCIA NAVEGADOR
            #navegador = webdriver.Chrome(options=chrome_options)
            navegador = webdriver.Edge(options=config_edge)



            # DEFININDO VARIAVEIS
            tipo_arquivo = "hapvida_boleto"



            # FLUXO DE NAVEGACAO

            # Acessar link boleto hapvida
            navegador.get("https://webhap.hapvida.com.br/pls/webhap/webNewBoletoEmpresa.Login")

            # Tempo para carregar a página
            time.sleep(5)

            # Fechar pop-up tela de login
            botao = navegador.find_element(By.CLASS_NAME, "ui-button-text")
            botao.click()

            time.sleep(1)
            navegador.maximize_window()

            # REALIZA LOGIN USANDO FUNCAO DE FORA (realizarLogin())
            realizarLogin(navegador, contrato, tipo_arquivo, "entidade", "cnpj")

            try:
                #Aguarda a tela do iframe entrar
                iframe_faturas = WebDriverWait(navegador, 20).until(
                    EC.presence_of_element_located((By.TAG_NAME, "iframe"))
                )
                
                # 2. Encontrou iframe, usar o switch_to.frame para acessá-lo
                # para caso precisar sair do iframe: navegador.switch_to.default_content()
                navegador.switch_to.frame(iframe_faturas)
    
            
                time.sleep(2)
                

            except Exception as e:
                print(f"Erro durante a execução: {e}")
                

            # Checa se existe alguma fatura, caso contrário, retorna ao for
            lista_avisos = navegador.find_elements(By.XPATH, "//h2[contains(string(), 'Nenhuma fatura encontrada')]")
            if len(lista_avisos) > 0:
                navegador.quit()
                print(f'{contrato} não consta na operadora!')
                continue
            
            print(contrato)
            

            try:
                wait = WebDriverWait(navegador, 60)
                linhas = wait.until(
                    EC.presence_of_all_elements_located((By.XPATH, "//div[contains(@class, '__tableRow')]"))
                )
                
                print(f'QUANTIDADE DE LINHAS: {len(linhas)}')
                cont = 0
                for linha in linhas:
                    cont += 1
                    print(cont)
                    try: 
                        valida_tipo_contrato = linha.find_element(
                        By.XPATH, f"./div[2][contains(string(), '{tipo_contrato_site}')]"
                        )

                        valida_venc_contrato = linha.find_element(
                        By.XPATH, f"./div[3][contains(string(), '{data_completa_venc}')]"
                        )


                        time.sleep(3)
                        abas_antigas = navegador.window_handles
                        
                        wait_linha = WebDriverWait(linha, 60)
                        botao_boleto =  wait_linha.until(
                            #linha.find_element(By.XPATH, ".//button[@type='button']").click()
                            EC.element_to_be_clickable((By.XPATH, ".//button[@type='button']"))
                        )
                        #Forçar click do botão
                        navegador.execute_script("arguments[0].click();", botao_boleto)
                        time.sleep(3)

                        # Aba do boleto
                        # Guarda todas as abas abertas antes do clique
                        
                        wait.until(EC.new_window_is_opened(abas_antigas))
            
                        todas_abas = navegador.window_handles
                        navegador.switch_to.window(todas_abas[-1])

                        # --- COMANDO FINAL PARA SALVAR ---

                        # Limpa qualquer arquivo que esteja na pasta download_principal
                        for f in os.listdir(pasta_download):
                            caminho_completo = os.path.join(pasta_download, f)
                            if os.path.isfile(caminho_completo):
                                os.remove(caminho_completo)
                        # Como foi ativado o Modo Kiosk, ele salva o PDF na pasta sem abrir a janela de impressão
                        
                        navegador.execute_script("window.print();")   
                        time.sleep(2)
                        # --- RENOMEANDO ARQUIVO DA PASTA DOWNLOAD ENCAMINHANDO PARA ESTRUTURA DE PASTAS ARQUIVOS_DIRECIONADOS ---

                        # Lista todos os arquivos PDF na pasta

                        pdfs = []

                        for f in os.listdir(pasta_download):
                            if f.endswith(".pdf"):
                                pdfs.append(f)

                        # Se houver pelo menos um PDF, renomeia o primeiro
                        if pdfs:
                            arquivo_antigo = os.path.join(pasta_download, pdfs[0])
                            arquivo_novo = os.path.join(pasta_download, f"{dia_venc}.{mes_venc}.{ano_venc}-{operadora} HAPVIDA {contrato} Boleto {tipo_contrato}.pdf")
                            os.rename(arquivo_antigo, arquivo_novo)
                            #print("Arquivo renomeado com sucesso!")
                        else:
                            print("Nenhum arquivo PDF encontrado.")
                            #exit()
                            continue


                        # Mover arquivo renomeado para a pasta de vencimento da operadora (Hapvida Arquivo)
                        shutil.move(arquivo_novo, pasta_destino)
                        print(f"BAIXADO: {arquivo_novo}")
                        # Fim da operacao
                        navegador.close()
                        time.sleep(1)
                        navegador.switch_to.window(todas_abas[0])
                        navegador.quit()
                        break


                    except NoSuchElementException:
                        if cont == len(linhas):
                            navegador.quit()
                        continue
            except TimeoutException:
                break
                
                






                